from django.http import JsonResponse
from django.views import View
from openpyxl import Workbook
from django.http import HttpResponse
from datetime import datetime, timedelta
from django.db.models import Count
from django.views import View

from .models import Robot
import json
from django.utils.dateparse import parse_datetime

# from django.utils.decorators import method_decorator
# from django.views.decorators.csrf import csrf_exempt

# Create your views here.


# @method_decorator(csrf_exempt, name='dispatch') - # убираем csrf для проверки
class AddRobotView(View):
    def post(self, request, *args, **kwargs):
        try:
            # оолучаем данные из тела запроса
            data = json.loads(request.body)
            model = data.get('model')
            version = data.get('version')
            created = data.get('created')

            # проверяем валидность модели
            if not model or len(model) != 2:
                return JsonResponse({'error': 'Некорректная модель'}, status=400)

            # проверяем валидность версии
            if not version or len(version) != 2:
                return JsonResponse({'error': 'Некорректная версия'}, status=400)

            # проверяем наличие даты создания
            if not created:
                return JsonResponse({'error': 'Дата создания не указана'}, status=400)

            # преобразуем строку в объект datetime
            created_datetime = parse_datetime(created)
            if not created_datetime:
                return JsonResponse({'error': 'Некорректный формат даты'}, status=400)

            # Формируем серийный номер
            serial = f"{model}-{version}"

            robot = Robot.objects.create(
                serial=serial,
                model=model,
                version=version,
                created=created_datetime
            )

            return JsonResponse({'message': f'Робот {robot.serial} успешно добавлен'}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Некорректный формат JSON'}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class RobotsExcelView(View):
    def get(self, request, *args, **kwargs):
        # определяем текущую дату и дату неделю назад
        today = datetime.today()
        week_ago = today - timedelta(days=7)

        # создаем новый Excel-файл
        workbook = Workbook()

        # получаем уникальные модели роботов
        robot_models = Robot.objects.values_list('model', flat=True).distinct()

        for model in robot_models:
            # создаем отдельный лист для каждой модели
            worksheet = workbook.create_sheet(title=model)
            worksheet.append(["Модель", "Версия", "Количество за неделю"])

            # считаем количество роботов каждой версии за последнюю неделю
            versions = Robot.objects.filter(
                model=model, created__gte=week_ago
            ).values('version').annotate(count=Count('id'))

            # записываем данные в лист
            for version_data in versions:
                worksheet.append([model, version_data['version'], version_data['count']])

        # удаляем стандартный лист, если он остался
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="robots_summary.xlsx"'

        workbook.save(response)
        return response
